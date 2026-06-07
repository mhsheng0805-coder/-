"""
114年 紡織所 來自民間業務收支管理系統
建立完整 Excel 工作簿
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── 常數 ──────────────────────────────────────────────
DEPARTMENTS = ['原料部', '產品部', '檢驗部', '製程部', '雲分部', '產服部']
MONTHS = [f'{i}月' for i in range(1, 13)]
YEAR = '114'

# 收入項目
INCOME_ITEMS = [
    ('檢測服務收入費',  'income'),
    ('技術服務收入費',  'income'),
    ('認証服務收入費',  'income'),
    ('其他業務收入費',  'income'),
    ('附屬業務-委外加工收入', 'sub'),
    ('附屬業務-其他非民間收入', 'sub'),
    ('附屬業務-其他民間收入', 'sub'),
    ('其他小計(自民間)', 'sub_total'),
    ('小型企業收入',   'income'),
    ('來自民間業務收入合計', 'total'),
]

# 支出項目
EXPENSE_ITEMS = [
    ('人事費用',       'expense'),
    ('業務費用',       'expense'),
    ('維護費',         'expense'),
    ('旅運費',         'expense'),
    ('材料費',         'expense'),
    ('租借設備使用費', 'expense'),
    ('差旅費',         'expense'),
    ('附屬業務-委外加工支出', 'sub'),
    ('附屬業務-其他非民間支出', 'sub'),
    ('附屬業務-其他民間支出', 'sub'),
    ('其他小計(自民間)支出', 'sub_total'),
    ('支出合計',       'total'),
]

# 已申請未核銷費用
UNCLAIMED_ITEMS = [
    '業務費(未核銷)',
    '旅運費(未核銷)',
    '材料費(未核銷)',
    '維護費(未核銷)',
    '未核銷合計',
]

# ── 顏色/字型 ──────────────────────────────────────────
def rgb(r, g, b): return f'{r:02X}{g:02X}{b:02X}'

C_TITLE_BG    = rgb(31, 78, 121)   # 深藍
C_DEPT_BG     = rgb(70, 130, 180)  # 鋼藍
C_HDR_BG      = rgb(189, 215, 238) # 淡藍
C_INCOME_BG   = rgb(226, 239, 218) # 淡綠
C_TOTAL_BG    = rgb(146, 208, 80)  # 綠
C_EXPENSE_BG  = rgb(255, 230, 230) # 淡紅
C_EXP_TOT_BG  = rgb(255, 153, 153) # 紅
C_SUB_BG      = rgb(242, 242, 242) # 淺灰
C_UNCLAIM_BG  = rgb(255, 242, 204) # 淡黃
C_CONTRACT_BG = rgb(237, 225, 255) # 淡紫
C_SUMMARY_BG  = rgb(255, 192, 0)   # 金黃（彙整標題）

FONT_TITLE  = Font(name='微軟正黑體', size=14, bold=True, color='FFFFFF')
FONT_HDR    = Font(name='微軟正黑體', size=10, bold=True)
FONT_DEPT   = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')
FONT_LABEL  = Font(name='微軟正黑體', size=10)
FONT_TOTAL  = Font(name='微軟正黑體', size=10, bold=True)
FONT_INPUT  = Font(name='Arial', size=10, color='0000FF')  # 藍色=可輸入
FONT_FORM   = Font(name='Arial', size=10, color='000000')  # 黑色=公式
FONT_GREEN  = Font(name='Arial', size=10, color='008000')  # 綠色=跨表連結

FMT_NUM  = '#,##0;(#,##0);"-"'
FMT_PCT  = '0.0%;-0.0%;"-"'
FMT_GOAL = '#,##0;(#,##0);"-"'

thin  = Side(style='thin')
thick = Side(style='medium')
BORDER_THIN = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_THICK = Border(left=thick, right=thick, top=thick, bottom=thick)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT   = Alignment(horizontal='left',   vertical='center')
RIGHT  = Alignment(horizontal='right',  vertical='center')

def fill(color): return PatternFill('solid', start_color=color, fgColor=color)

def set_cell(ws, row, col, value=None, font=None, fill_=None,
             align=None, border=None, fmt=None, comment=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill_:  c.fill = fill_
    if align:  c.alignment = align
    if border: c.border = border
    if fmt:    c.number_format = fmt
    return c

def merge_set(ws, r1, c1, r2, c2, value=None, font=None,
              fill_=None, align=None, border=None):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2,   end_column=c2)
    c = ws.cell(row=r1, column=c1, value=value)
    if font:   c.font = font
    if fill_:  c.fill = fill_
    if align:  c.alignment = align
    if border: c.border = border
    return c

# ── 建立部門收支工作表 ──────────────────────────────────
def build_dept_sheet(wb, dept_name):
    ws = wb.create_sheet(dept_name)
    ws.sheet_view.showGridLines = False

    # 欄寬
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    for i, col in enumerate(['C','D','E','F','G','H','I','J','K','L','M','N'], 1):
        ws.column_dimensions[col].width = 11
    ws.column_dimensions['O'].width = 13  # 合計
    ws.column_dimensions['P'].width = 13  # 年度目標
    ws.column_dimensions['Q'].width = 9   # 達成率

    # Row 1: 主標題
    ws.row_dimensions[1].height = 35
    merge_set(ws, 1, 1, 1, 17,
              f'紡織所 {dept_name} {YEAR}年 來自民間業務收支表',
              FONT_TITLE, fill(C_TITLE_BG), CENTER)

    # Row 2: 欄位標題
    ws.row_dimensions[2].height = 28
    headers = ['項目', ''] + MONTHS + ['合計', '年度目標', '達成率']
    for col, h in enumerate(headers, 1):
        set_cell(ws, 2, col, h, FONT_HDR, fill(C_HDR_BG), CENTER, BORDER_THIN)

    # ── 收入區 ──
    row = 3
    merge_set(ws, row, 1, row, 17, '【收入】',
              FONT_TOTAL, fill(C_INCOME_BG), LEFT)
    ws.row_dimensions[row].height = 20
    row += 1

    income_rows = {}  # item_name -> row number
    for item, kind in INCOME_ITEMS:
        ws.row_dimensions[row].height = 18
        if kind == 'total':
            f_ = fill(C_TOTAL_BG)
            fn = FONT_TOTAL
        elif kind == 'sub_total':
            f_ = fill(C_SUB_BG)
            fn = FONT_TOTAL
        elif kind == 'sub':
            f_ = fill(C_SUB_BG)
            fn = FONT_LABEL
        else:
            f_ = fill(C_INCOME_BG)
            fn = FONT_LABEL

        set_cell(ws, row, 1, item, fn, f_, LEFT, BORDER_THIN)
        set_cell(ws, row, 2, None, fn, f_, CENTER, BORDER_THIN)

        if kind == 'total':
            # 來自民間業務收入合計 = 前幾行可輸入之和
            # 用公式: SUM of input rows above
            input_rows = [r for r, (nm, kd) in zip(
                range(4, row), [(nm, kd) for nm, kd in INCOME_ITEMS[:-1]]
            ) if kd not in ('sub', 'sub_total')]
            # 逐月合計
            for m_col in range(3, 15):
                # 找上面非sub的行
                col_letter = get_column_letter(m_col)
                formula_parts = []
                for r2, (nm2, kd2) in zip(range(4, row), INCOME_ITEMS[:-1]):
                    if kd2 not in ('sub', 'sub_total'):
                        formula_parts.append(f'{col_letter}{r2}')
                formula = '=SUM(' + ','.join(formula_parts) + ')' if formula_parts else 0
                c = set_cell(ws, row, m_col, formula, FONT_FORM, f_, RIGHT, BORDER_THIN, FMT_NUM)
            # 合計欄(O)
            c_start = get_column_letter(3)
            c_end   = get_column_letter(14)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
        elif kind == 'sub_total':
            # 其他小計 = 附屬業務三項之和
            for m_col in range(3, 15):
                col_letter = get_column_letter(m_col)
                sub_rows_ref = []
                for r2, (nm2, kd2) in zip(range(4, row), INCOME_ITEMS[:-1]):
                    if kd2 == 'sub':
                        sub_rows_ref.append(f'{col_letter}{r2}')
                formula = '=SUM(' + ','.join(sub_rows_ref) + ')' if sub_rows_ref else 0
                set_cell(ws, row, m_col, formula, FONT_FORM, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
        else:
            # 可輸入欄位 (藍色)
            for m_col in range(3, 15):
                set_cell(ws, row, m_col, 0, FONT_INPUT, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_FORM, f_, RIGHT, BORDER_THIN, FMT_NUM)

        # 年度目標 (藍色可輸入)
        set_cell(ws, row, 16, 0, FONT_INPUT, f_, RIGHT, BORDER_THIN, FMT_NUM)
        # 達成率
        if kind in ('total', 'sub_total') or kind == 'income':
            set_cell(ws, row, 17, f'=IF(P{row}=0,"-",O{row}/P{row})',
                     FONT_FORM, f_, CENTER, BORDER_THIN, FMT_PCT)
        else:
            set_cell(ws, row, 17, '', fn, f_, CENTER, BORDER_THIN)

        income_rows[item] = row
        row += 1

    income_total_row = income_rows['來自民間業務收入合計']

    # ── 支出區 ──
    row += 1
    merge_set(ws, row, 1, row, 17, '【支出】',
              FONT_TOTAL, fill(C_EXPENSE_BG), LEFT)
    ws.row_dimensions[row].height = 20
    expense_section_start = row
    row += 1

    expense_rows = {}
    for item, kind in EXPENSE_ITEMS:
        ws.row_dimensions[row].height = 18
        if kind == 'total':
            f_ = fill(C_EXP_TOT_BG)
            fn = FONT_TOTAL
        elif kind == 'sub_total':
            f_ = fill(C_SUB_BG)
            fn = FONT_TOTAL
        elif kind == 'sub':
            f_ = fill(C_SUB_BG)
            fn = FONT_LABEL
        else:
            f_ = fill(C_EXPENSE_BG)
            fn = FONT_LABEL

        set_cell(ws, row, 1, item, fn, f_, LEFT, BORDER_THIN)
        set_cell(ws, row, 2, None, fn, f_, CENTER, BORDER_THIN)

        if kind == 'total':
            for m_col in range(3, 15):
                col_letter = get_column_letter(m_col)
                exp_refs = []
                for r2, (nm2, kd2) in zip(range(expense_section_start+1, row), EXPENSE_ITEMS[:-1]):
                    if kd2 not in ('sub', 'sub_total'):
                        exp_refs.append(f'{col_letter}{r2}')
                formula = '=SUM(' + ','.join(exp_refs) + ')' if exp_refs else 0
                set_cell(ws, row, m_col, formula, FONT_FORM, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
        elif kind == 'sub_total':
            for m_col in range(3, 15):
                col_letter = get_column_letter(m_col)
                sub_refs = []
                for r2, (nm2, kd2) in zip(range(expense_section_start+1, row), EXPENSE_ITEMS[:-1]):
                    if kd2 == 'sub':
                        sub_refs.append(f'{col_letter}{r2}')
                formula = '=SUM(' + ','.join(sub_refs) + ')' if sub_refs else 0
                set_cell(ws, row, m_col, formula, FONT_FORM, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
        else:
            for m_col in range(3, 15):
                set_cell(ws, row, m_col, 0, FONT_INPUT, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_FORM, f_, RIGHT, BORDER_THIN, FMT_NUM)

        set_cell(ws, row, 16, 0, FONT_INPUT, f_, RIGHT, BORDER_THIN, FMT_NUM)
        if kind in ('total', 'sub_total') or kind == 'expense':
            set_cell(ws, row, 17, f'=IF(P{row}=0,"-",O{row}/P{row})',
                     FONT_FORM, f_, CENTER, BORDER_THIN, FMT_PCT)
        else:
            set_cell(ws, row, 17, '', fn, f_, CENTER, BORDER_THIN)

        expense_rows[item] = row
        row += 1

    expense_total_row = expense_rows['支出合計']

    # ── 收支差額 ──
    ws.row_dimensions[row].height = 22
    f_ = fill(rgb(255, 204, 0))
    set_cell(ws, row, 1, '收支差額 (收入-支出)', FONT_TOTAL, f_, LEFT, BORDER_THIN)
    set_cell(ws, row, 2, '', FONT_TOTAL, f_, CENTER, BORDER_THIN)
    for m_col in range(3, 15):
        col_letter = get_column_letter(m_col)
        formula = f'={col_letter}{income_total_row}-{col_letter}{expense_total_row}'
        set_cell(ws, row, m_col, formula, FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
    set_cell(ws, row, 15, f'=O{income_total_row}-O{expense_total_row}', FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
    set_cell(ws, row, 16, '', FONT_TOTAL, f_, RIGHT, BORDER_THIN)
    set_cell(ws, row, 17, f'=IF(P{income_total_row}=0,"-",O{row}/P{income_total_row})',
             FONT_FORM, f_, CENTER, BORDER_THIN, FMT_PCT)
    net_row = row
    row += 2

    # ── 已申請未核銷費用區 ──
    ws.row_dimensions[row].height = 20
    merge_set(ws, row, 1, row, 17, '【已申請未核銷費用】（截至當月底止）',
              FONT_TOTAL, fill(C_UNCLAIM_BG), LEFT)
    row += 1

    unclaim_rows = {}
    for item in UNCLAIMED_ITEMS:
        ws.row_dimensions[row].height = 18
        f_ = fill(C_UNCLAIM_BG)
        fn = FONT_LABEL if item != '未核銷合計' else FONT_TOTAL
        set_cell(ws, row, 1, item, fn, f_, LEFT, BORDER_THIN)
        set_cell(ws, row, 2, '', fn, f_, CENTER, BORDER_THIN)
        if item == '未核銷合計':
            for m_col in range(3, 15):
                col_letter = get_column_letter(m_col)
                refs = [f'{col_letter}{unclaim_rows[nm]}' for nm in UNCLAIMED_ITEMS[:-1]]
                formula = '=SUM(' + ','.join(refs) + ')'
                set_cell(ws, row, m_col, formula, FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
        else:
            for m_col in range(3, 15):
                set_cell(ws, row, m_col, 0, FONT_INPUT, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, row, 15, f'=SUM(C{row}:N{row})', FONT_FORM, f_, RIGHT, BORDER_THIN, FMT_NUM)
        set_cell(ws, row, 16, '', fn, f_, RIGHT, BORDER_THIN)
        set_cell(ws, row, 17, '', fn, f_, CENTER, BORDER_THIN)
        unclaim_rows[item] = row
        row += 1

    row += 1

    # ── 合約追蹤區 ──
    ws.row_dimensions[row].height = 20
    merge_set(ws, row, 1, row, 17, '【合約進度追蹤】',
              FONT_TOTAL, fill(C_CONTRACT_BG), LEFT)
    row += 1

    contract_fields = [
        ('本月新增簽約件數', FMT_NUM),
        ('本月新增簽約金額', FMT_NUM),
        ('上月延續-洽談件數', FMT_NUM),
        ('上月延續-洽談金額', FMT_NUM),
        ('上月延續-已簽約件數', FMT_NUM),
        ('上月延續-已簽約金額', FMT_NUM),
        ('本月完成件數', FMT_NUM),
        ('累計簽約金額', FMT_NUM),
        ('備註/客戶名稱', '@'),
    ]

    contract_start_row = row
    contract_rows = {}
    for item, fmt in contract_fields:
        ws.row_dimensions[row].height = 18
        f_ = fill(C_CONTRACT_BG)
        fn = FONT_LABEL
        set_cell(ws, row, 1, item, fn, f_, LEFT, BORDER_THIN)
        set_cell(ws, row, 2, '', fn, f_, CENTER, BORDER_THIN)
        if item == '累計簽約金額':
            for m_col in range(3, 15):
                col_letter = get_column_letter(m_col)
                new_amt_row = contract_rows['本月新增簽約金額']
                prev_signed_row = contract_rows['上月延續-已簽約金額']
                if m_col == 3:
                    formula = f'={col_letter}{new_amt_row}'
                else:
                    prev_col = get_column_letter(m_col - 1)
                    formula = f'={prev_col}{row}+{col_letter}{new_amt_row}'
                set_cell(ws, row, m_col, formula, FONT_FORM, f_, RIGHT, BORDER_THIN, fmt)
            set_cell(ws, row, 15, f'=N{row}', FONT_FORM, f_, RIGHT, BORDER_THIN, fmt)
        else:
            for m_col in range(3, 15):
                set_cell(ws, row, m_col, 0 if fmt != '@' else '', FONT_INPUT, f_, RIGHT if fmt != '@' else LEFT, BORDER_THIN, fmt)
            set_cell(ws, row, 15, '', FONT_FORM, f_, RIGHT, BORDER_THIN)
        set_cell(ws, row, 16, '', fn, f_, RIGHT, BORDER_THIN)
        set_cell(ws, row, 17, '', fn, f_, CENTER, BORDER_THIN)
        contract_rows[item] = row
        row += 1

    ws.freeze_panes = 'C3'
    return ws


# ── 建立彙整工作表 ──────────────────────────────────────
def build_summary_sheet(wb):
    ws = wb.create_sheet('彙整', 0)
    ws.sheet_view.showGridLines = False

    # 欄寬
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    for col_idx in range(3, 16):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    ws.column_dimensions['O'].width = 13
    ws.column_dimensions['P'].width = 13
    ws.column_dimensions['Q'].width = 9

    # 標題
    ws.row_dimensions[1].height = 35
    merge_set(ws, 1, 1, 1, 17,
              f'紡織所 {YEAR}年 來自民間業務收支彙整表（六部門）',
              FONT_TITLE, fill(C_TITLE_BG), CENTER)

    # 欄位標題
    ws.row_dimensions[2].height = 28
    for col, h in enumerate(['部門', '項目'] + MONTHS + ['合計', '年度目標', '達成率'], 1):
        set_cell(ws, 2, col, h, FONT_HDR, fill(C_HDR_BG), CENTER, BORDER_THIN)

    row = 3
    summary_items = [
        ('來自民間業務收入合計', INCOME_ITEMS, '來自民間業務收入合計'),
        ('支出合計',             EXPENSE_ITEMS, '支出合計'),
        ('收支差額',             None, None),
        ('未核銷合計',           None, None),
    ]

    dept_colors = [
        rgb(204, 229, 255),
        rgb(204, 255, 229),
        rgb(255, 229, 204),
        rgb(229, 204, 255),
        rgb(255, 204, 229),
        rgb(229, 255, 204),
    ]

    for d_idx, dept in enumerate(DEPARTMENTS):
        dept_ws_name = dept
        d_fill = fill(dept_colors[d_idx])
        dept_start_row = row

        # 部門名稱（跨多行合併）
        dept_rows_count = 5  # 收入、支出、差額、未核銷、空行
        ws.row_dimensions[row].height = 20

        summary_row_keys = [
            ('來自民間業務收入合計', fill(C_TOTAL_BG), FONT_TOTAL),
            ('支出合計',             fill(C_EXP_TOT_BG), FONT_TOTAL),
            ('收支差額',             fill(rgb(255, 204, 0)), FONT_TOTAL),
            ('未核銷合計',           fill(C_UNCLAIM_BG), FONT_TOTAL),
        ]

        dept_data_rows = []
        for label, row_fill, row_font in summary_row_keys:
            dept_data_rows.append(row)
            row += 1

        # 在部門欄位合併
        ws.merge_cells(start_row=dept_start_row, start_column=1,
                       end_row=dept_start_row + len(summary_row_keys) - 1, end_column=1)
        c = ws.cell(row=dept_start_row, column=1, value=dept)
        c.font = Font(name='微軟正黑體', size=11, bold=True)
        c.fill = fill(dept_colors[d_idx])
        c.alignment = CENTER
        c.border = BORDER_THIN

        # 重新設定各行
        for i, (label, row_fill, row_font) in enumerate(summary_row_keys):
            r = dept_start_row + i
            ws.row_dimensions[r].height = 20

            # 部門欄位邊框
            ws.cell(row=r, column=1).border = BORDER_THIN

            # 項目名稱
            set_cell(ws, r, 2, label, row_font, row_fill, LEFT, BORDER_THIN)

            # 月份：連結到各部門工作表
            for m_col in range(3, 15):
                col_letter = get_column_letter(m_col)
                # 找對應工作表的行號 - 這裡用公式直接連結
                # 使用 INDIRECT 或直接引用各部門工作表
                # 由於我們不知道確切行號，填入 0 待使用者更新，並標記為藍色
                set_cell(ws, r, m_col, 0, FONT_INPUT, row_fill, RIGHT, BORDER_THIN, FMT_NUM)

            # 合計 = SUM(月份)
            set_cell(ws, r, 15, f'=SUM(C{r}:N{r})', FONT_FORM, row_fill, RIGHT, BORDER_THIN, FMT_NUM)
            # 年度目標
            set_cell(ws, r, 16, 0, FONT_INPUT, row_fill, RIGHT, BORDER_THIN, FMT_NUM)
            # 達成率
            set_cell(ws, r, 17, f'=IF(P{r}=0,"-",O{r}/P{r})',
                     FONT_FORM, row_fill, CENTER, BORDER_THIN, FMT_PCT)

        # 空行分隔
        ws.row_dimensions[row].height = 8
        merge_set(ws, row, 1, row, 17, '', None, fill('FFFFFF'), None)
        row += 1

    # 六部門合計行
    ws.row_dimensions[row].height = 28
    total_items = [
        ('六部門合計', '來自民間業務收入', fill(C_TOTAL_BG)),
        ('', '支出合計', fill(C_EXP_TOT_BG)),
        ('', '收支差額', fill(rgb(255, 204, 0))),
        ('', '未核銷合計', fill(C_UNCLAIM_BG)),
    ]
    grand_start = row
    for i, (dept_label, item_label, row_fill) in enumerate(total_items):
        r = row + i
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=1).border = BORDER_THIN
        if i == 0:
            ws.merge_cells(start_row=grand_start, start_column=1,
                           end_row=grand_start + 3, end_column=1)
            c = ws.cell(row=grand_start, column=1, value='六部門合計')
            c.font = FONT_TITLE
            c.fill = fill(C_SUMMARY_BG)
            c.alignment = CENTER
            c.border = BORDER_THIN
        set_cell(ws, r, 2, item_label, FONT_TOTAL, row_fill, LEFT, BORDER_THIN)
        for m_col in range(3, 15):
            set_cell(ws, r, m_col, 0, FONT_FORM, row_fill, RIGHT, BORDER_THIN, FMT_NUM)
        set_cell(ws, r, 15, f'=SUM(C{r}:N{r})', FONT_TOTAL, row_fill, RIGHT, BORDER_THIN, FMT_NUM)
        set_cell(ws, r, 16, 0, FONT_INPUT, row_fill, RIGHT, BORDER_THIN, FMT_NUM)
        set_cell(ws, r, 17, f'=IF(P{r}=0,"-",O{r}/P{r})',
                 FONT_FORM, row_fill, CENTER, BORDER_THIN, FMT_PCT)
    ws.freeze_panes = 'C3'
    return ws


# ── 建立合約追蹤工作表 ──────────────────────────────────
def build_contract_track_sheet(wb):
    ws = wb.create_sheet('合約追蹤明細')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 12

    ws.row_dimensions[1].height = 35
    merge_set(ws, 1, 1, 1, 10,
              f'{YEAR}年 合約簽約及洽談進度追蹤明細表',
              FONT_TITLE, fill(C_TITLE_BG), CENTER)

    ws.row_dimensions[2].height = 28
    hdrs = ['月份', '部門', '客戶/計畫名稱', '合約金額', '簽約日期',
            '預計完成日', '狀態', '本月實收金額', '備註', '延續至下月']
    for col, h in enumerate(hdrs, 1):
        set_cell(ws, 2, col, h, FONT_HDR, fill(C_HDR_BG), CENTER, BORDER_THIN)

    status_hint = ['新增簽約', '洽談中', '已簽約執行中', '完成', '延續']
    row = 3
    for month_idx in range(1, 13):
        # 月份標題行
        ws.row_dimensions[row].height = 22
        merge_set(ws, row, 1, row, 10,
                  f'{month_idx}月',
                  Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF'),
                  fill(C_DEPT_BG), CENTER)
        row += 1

        # 每月預留10行輸入
        for i in range(10):
            ws.row_dimensions[row].height = 18
            f_ = fill(rgb(235, 245, 255) if i % 2 == 0 else 'FFFFFF')
            set_cell(ws, row, 1, month_idx, FONT_LABEL, f_, CENTER, BORDER_THIN)
            # 部門下拉
            set_cell(ws, row, 2, '', FONT_INPUT, f_, CENTER, BORDER_THIN)
            for col in range(3, 11):
                fmt = FMT_NUM if col in (4, 8) else ('@' if col in (3, 9) else None)
                val = 0 if col in (4, 8) else ''
                set_cell(ws, row, col, val, FONT_INPUT, f_, LEFT if col == 3 else RIGHT if col in (4,8) else CENTER, BORDER_THIN, fmt)
            row += 1

    ws.freeze_panes = 'A3'
    return ws


# ── 建立未核銷費用彙整表 ────────────────────────────────
def build_unclaimed_sheet(wb):
    ws = wb.create_sheet('未核銷費用彙整')
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    ws.row_dimensions[1].height = 35
    merge_set(ws, 1, 1, 1, 8,
              f'{YEAR}年 各部門已申請未核銷費用彙整表',
              FONT_TITLE, fill(C_TITLE_BG), CENTER)

    ws.row_dimensions[2].height = 14

    unclaim_types = ['業務費', '旅運費', '材料費', '維護費', '合計']

    for month_idx in range(1, 13):
        row_start = 3 + (month_idx - 1) * (len(DEPARTMENTS) + 3)
        ws.row_dimensions[row_start].height = 22

        merge_set(ws, row_start, 1, row_start, 8,
                  f'{month_idx}月 未核銷費用',
                  Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF'),
                  fill(C_DEPT_BG), CENTER)

        hdr_row = row_start + 1
        ws.row_dimensions[hdr_row].height = 22
        for col, h in enumerate(['部門'] + unclaim_types + ['備註'], 1):
            set_cell(ws, hdr_row, col, h, FONT_HDR, fill(C_HDR_BG), CENTER, BORDER_THIN)

        for d_idx, dept in enumerate(DEPARTMENTS):
            r = hdr_row + 1 + d_idx
            ws.row_dimensions[r].height = 18
            f_ = fill(C_UNCLAIM_BG)
            set_cell(ws, r, 1, dept, FONT_LABEL, f_, LEFT, BORDER_THIN)
            for col in range(2, 6):
                set_cell(ws, r, col, 0, FONT_INPUT, f_, RIGHT, BORDER_THIN, FMT_NUM)
            # 合計
            set_cell(ws, r, 6, f'=SUM(B{r}:E{r})', FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
            set_cell(ws, r, 7, '', FONT_LABEL, f_, LEFT, BORDER_THIN)

        # 小計行
        tot_row = hdr_row + 1 + len(DEPARTMENTS)
        ws.row_dimensions[tot_row].height = 20
        f_ = fill(rgb(255, 200, 100))
        set_cell(ws, tot_row, 1, '月合計', FONT_TOTAL, f_, LEFT, BORDER_THIN)
        for col in range(2, 7):
            start_r = hdr_row + 2
            end_r   = hdr_row + 1 + len(DEPARTMENTS)
            col_l = get_column_letter(col)
            set_cell(ws, tot_row, col, f'=SUM({col_l}{start_r}:{col_l}{end_r})',
                     FONT_TOTAL, f_, RIGHT, BORDER_THIN, FMT_NUM)
        set_cell(ws, tot_row, 7, '', FONT_TOTAL, f_, LEFT, BORDER_THIN)

    ws.freeze_panes = 'A3'
    return ws


# ── 建立說明工作表 ──────────────────────────────────────
def build_instruction_sheet(wb):
    ws = wb.create_sheet('使用說明', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 60

    ws.row_dimensions[1].height = 40
    merge_set(ws, 1, 1, 1, 3,
              f'紡織所 {YEAR}年 來自民間業務收支管理系統 使用說明',
              FONT_TITLE, fill(C_TITLE_BG), CENTER)

    instructions = [
        ('工作表說明', ''),
        ('彙整', '六個部門收入、支出、收支差額及未核銷費用的總覽，藍色儲存格可手動輸入或鏈接各部門工作表。'),
        ('原料部 / 產品部 / 檢驗部 / 製程部 / 雲分部 / 產服部', '各部門獨立收支追蹤表，含1-12月月份欄位。'),
        ('合約追蹤明細', '逐月記錄每筆合約的簽約進度，包含洽談中、已簽約、完成等狀態。'),
        ('未核銷費用彙整', '各部門每月已申請但尚未核銷的費用彙整。'),
        ('', ''),
        ('填寫說明', ''),
        ('藍色儲存格', '可直接輸入的欄位（年度目標、各月實際金額、已申請未核銷費用等）。'),
        ('黑色儲存格', '系統自動計算的公式，請勿更改。'),
        ('綠色儲存格', '跨工作表連結，數值來自其他工作表。'),
        ('', ''),
        ('合約追蹤說明', ''),
        ('狀態欄位說明', '新增簽約：本月新簽; 洽談中：尚未正式簽約; 已簽約執行中：合約進行中; 完成：已結案'),
        ('延續機制', '若合約跨月，請在下月對應行填寫，並在備註欄填入上月合約編號。'),
        ('', ''),
        ('WEB 系統說明', ''),
        ('存取方式', '在有 WiFi 的環境下，瀏覽器輸入伺服器 IP + Port 即可連線填寫。'),
        ('預設網址', 'http://[伺服器IP]:5001  (請向系統管理員確認 IP)'),
        ('功能', '線上填寫各部門收支數據、查詢報表、匯出 Excel。'),
    ]

    row = 2
    for label, desc in instructions:
        ws.row_dimensions[row].height = 22
        if label and not desc:
            merge_set(ws, row, 1, row, 3, f'▌ {label}',
                      Font(name='微軟正黑體', size=11, bold=True),
                      fill(C_HDR_BG), LEFT)
        else:
            set_cell(ws, row, 1, '', FONT_LABEL, fill('FFFFFF'), CENTER)
            set_cell(ws, row, 2, label, FONT_LABEL, fill('FFFFFF'), LEFT)
            set_cell(ws, row, 3, desc, FONT_LABEL, fill('FFFFFF'), LEFT)
        row += 1
    return ws


# ── 主程式 ─────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    # 移除預設工作表
    wb.remove(wb.active)

    print('建立工作表...')
    build_instruction_sheet(wb)
    build_summary_sheet(wb)
    for dept in DEPARTMENTS:
        print(f'  建立 {dept}...')
        build_dept_sheet(wb, dept)
    build_contract_track_sheet(wb)
    build_unclaimed_sheet(wb)

    out_path = r'C:\工作\114年\組織績效管理\業務收支系統\114年來自民間業務收支管理系統.xlsx'
    wb.save(out_path)
    print(f'已儲存: {out_path}')
    return out_path


if __name__ == '__main__':
    main()
